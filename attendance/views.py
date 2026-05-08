# attendance/views.py

from datetime import datetime, timedelta, date
from calendar import monthrange
from io import BytesIO
from decimal import Decimal, ROUND_HALF_UP
from collections import defaultdict, OrderedDict

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.db import transaction

import qrcode
import openpyxl
from openpyxl.utils import get_column_letter

# ⬇️ NUEVOS IMPORTS (para kiosco/QR firmado)
from django.core.signing import TimestampSigner, SignatureExpired, BadSignature
from django.views.decorators.csrf import csrf_protect
from urllib.parse import urlencode, quote as urlquote

# Forms
from .forms import (
    EmployeeForm, ScanForm,
    PayrollPeriodForm, PayrollLineAdjustmentForm, PayrollExportForm,
    ProjectForm, IncomeForm, ExpenseForm
)

# Models
from .models import (
    Employee, Attendance, Project,
    PayrollPeriod, PayrollLine,
    ProjectIncome, ProjectExpense
)
import os
from django.conf import settings
from openpyxl.drawing.image import Image as XLImage

# ──────────────────────────────────────────────────────────────────────────────
# Helpers generales
# ──────────────────────────────────────────────────────────────────────────────
def _add_logo(ws, cell="A1", max_width_px=150):
    """
    Inserta el logo en la hoja, ajusta fila/columna para que no tape el título.
    Busca: /static/img/LOGO_PERFIL_AMBU_GROUP.png
    """
    logo_path = os.path.join(settings.BASE_DIR, "static", "img", "LOGO_PERFIL_AMBU_GROUP.png")
    if not os.path.exists(logo_path):
        return

    # Altura de fila 1 y ancho de columna A para dar espacio visual al logo
    ws.row_dimensions[1].height = 85  # ~85pt ≈ 113px aprox.
    ws.column_dimensions["A"].width = 23  # espacio al logo

    img = XLImage(logo_path)

    # Redimensiona manteniendo proporción (si hace falta)
    try:
        from PIL import Image as PILImage
        with PILImage.open(logo_path) as im:
            w, h = im.size
        if w > max_width_px:
            ratio = max_width_px / float(w)
            img.width = int(w * ratio)
            img.height = int(h * ratio)
    except Exception:
        # Si no hay Pillow, seguimos sin redimensionar
        pass

    ws.add_image(img, cell)

def _user_can_access_project(user, project: Project) -> bool:
    """Permite superuser o supervisor asignado al proyecto."""
    return user.is_superuser or project.supervisors.filter(id=user.id).exists()

def is_admin(user):
    return user.is_superuser

# ----- Formatos LATAM -----

def format_money_latam(value: Decimal | float | int) -> str:
    """$  1.225,00 con puntos de miles y coma decimal (2 dec)."""
    q = Decimal(value or 0).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    us = f"{q:,.2f}"                     # 1,234,567.89
    latam = us.replace(",", "§").replace(".", ",").replace("§", ".")
    return f"$  {latam}"

def format_money_latam_whole(value: Decimal | float | int) -> str:
    """$  1.226 con redondeo HALF-UP a entero para UI de nómina."""
    q = Decimal(value or 0).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    us = f"{q:,.0f}"
    latam = us.replace(",", "§").replace(".", ",").replace("§", ".")
    return f"$  {latam}"

def format_hhmm_from_minutes(minutes: int | float | Decimal) -> str:
    """Convierte minutos a HH:MM (solo visual)."""
    m = int(Decimal(minutes or 0))
    return f"{m//60:02d}:{m%60:02d}"

def format_hours_latam(value: Decimal | float | int) -> str:
    """8,5 (sin ,0 final). Redondeo a 0.5h típico de nómina."""
    q = Decimal(value or 0).quantize(Decimal("0.5"), rounding=ROUND_HALF_UP)
    s = f"{q:.1f}".replace(".", ",")
    return s[:-2] if s.endswith(",0") else s

def minutes_to_hours_decimal(minutes: int) -> Decimal:
    """Horas con 2 decimales (visual)."""
    return (Decimal(minutes) / Decimal(60)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

# ----- Parse HH:MM/decimales para overrides en Reporte -----

def _parse_hhmm_to_minutes(s: str) -> int:
    """
    Acepta 'HH:MM' o 'H:MM' o decimales como '8' o '8.5'.
    Devuelve minutos >= 0.
    """
    s = (s or "").strip()
    if not s:
        raise ValueError("Vacío")
    if ":" in s:
        hh, mm = s.split(":", 1)
        h = int(hh)
        m = int(mm)
        if not (0 <= m < 60):
            raise ValueError("Minutos inválidos")
        return max(0, h * 60 + m)
    q = Decimal(s)
    return max(0, int((q * Decimal(60)).to_integral_value(rounding=ROUND_HALF_UP)))

# ----- Descuento de almuerzo (1 hora) -----

def _billable_minutes(att: Attendance) -> int:
    """
    Minutos netos a pagar (según Attendance.worked_minutes()).
    No hacer ningún descuento adicional aquí.
    """
    try:
        return int(att.worked_minutes())   # ya descuenta 1h si > 5h
    except Exception:
        return int(getattr(att, "minutes", 0) or 0)


# ----- Semanas (LUN–SAB) dentro de un rango (usado en reportes) -----

DAYS_ORDER = ["LUNES", "MARTES", "MIERCOLES", "JUEVES", "VIERNES", "SABADO"]

def week_chunks_within(start: date, end: date):
    """
    Divide [start, end] en semanas LUN–SAB (recortadas al rango).
    Devuelve lista de (ini_sem, fin_sem).
    """
    start_monday = start - timedelta(days=start.weekday())  # lunes
    cursor = start_monday
    weeks = []
    while cursor <= end:
        week_start = cursor
        week_end = cursor + timedelta(days=5)  # lun->sab
        real_start = max(week_start, start)
        real_end = min(week_end, end)
        if real_start <= real_end:
            weeks.append((real_start, real_end))
        cursor += timedelta(days=7)
    return weeks

def biweekly_weeks(start: date, end: date):
    weeks = week_chunks_within(start, end)
    if len(weeks) > 2:
        weeks = weeks[:2]
    return weeks

def _employee_full_name(emp: Employee) -> str:
    return f"{emp.first_name} {emp.last_name}".strip()

def _build_week_table_report(employees, att_map, week_start: date, week_end: date):
    """
    Construye filas para una semana (reporte): horas por día, totales y pagos.
    att_map: dict[(emp_id, date)] -> horas (Decimal)
    """
    rows = []
    total_payroll = Decimal("0")
    total_hours_all = Decimal("0")

    # map day_name -> date real
    day_dates = OrderedDict()
    d = week_start
    while d <= week_end:
        if d.weekday() < 6:  # 0..5 = lun..sab
            day_dates[DAYS_ORDER[d.weekday()]] = d
        d += timedelta(days=1)

    for emp in employees:
        day_hours = []
        total_hours = Decimal("0")
        for _, dd in day_dates.items():
            h = att_map.get((emp.id, dd), Decimal("0"))
            day_hours.append(h)
            total_hours += h

        hourly = emp.hourly_rate or Decimal("0")
        total_pay = (total_hours * hourly).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        rows.append({
            "name": _employee_full_name(emp),
            "day_hours_raw": day_hours,
            "day_hours": [format_hours_latam(h) for h in day_hours],
            "total_hours_raw": total_hours,
            "total_hours": format_hours_latam(total_hours),
            "hourly_raw": hourly,
            "hourly": format_money_latam(hourly),
            "total_pay_raw": total_pay,
            "total_pay": format_money_latam(total_pay),
        })

        total_payroll += total_pay
        total_hours_all += total_hours

    return {
        "week_start": week_start,
        "week_end": week_end,
        "columns": list(day_dates.keys()),   # ["LUNES",..., "SABADO"]
        "rows": rows,
        "total_hours_all_raw": total_hours_all,
        "total_hours_all": format_hours_latam(total_hours_all),
        "total_payroll_raw": total_payroll,
        "total_payroll": format_money_latam(total_payroll),
        "show_divider": True,
    }

# ──────────────────────────────────────────────────────────────────────────────
# Dashboard
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def dashboard_view(request):
    """Lista proyectos según permisos del usuario."""
    if request.user.is_superuser:
        projects = Project.objects.all()
    else:
        projects = request.user.supervised_projects.all()
    return render(request, "attendance/dashboard.html", {"projects": projects})

# ──────────────────────────────────────────────────────────────────────────────
# Scan (registro por proyecto) - versión interna (requiere login)
# ──────────────────────────────────────────────────────────────────────────────

from django.http import JsonResponse
from django.db import transaction, IntegrityError
import uuid

@transaction.atomic
@login_required
def scan_view(request, project_slug):
    project = get_object_or_404(Project, slug=project_slug)
    if not _user_can_access_project(request.user, project):
        raise Http404("Proyecto no disponible para tu usuario")

    if request.method == "POST":
        form = ScanForm(request.POST, request.FILES, project=project)
        if not form.is_valid():
            # Si el cliente espera JSON, devolvemos error explícito
            if request.headers.get("Accept", "").find("application/json") >= 0:
                return JsonResponse({"ok": False, "error": "Formulario inválido"}, status=400)
            messages.error(request, "Datos inválidos.")
            return redirect("scan", project_slug=project.slug)

        employee = form.cleaned_data["employee"]
        action   = form.cleaned_data["action"]
        photo    = form.cleaned_data["evidence_photo"]  # obligatorio

        # Idempotencia desde el cliente (opcional pero recomendado)
        client_uuid = request.POST.get("client_uuid") or str(uuid.uuid4())
        device_ts   = request.POST.get("device_ts")     # ISO opcional

        today = timezone.localdate()
        now   = timezone.now()

        # Garantizar vínculo empleado↔proyecto
        project.employees.add(employee)

        # 1) Registrar escaneo idempotente
        try:
            from .models import AttendanceScan  # evita import circular
            AttendanceScan.objects.create(
                client_uuid=client_uuid,
                employee=employee,
                project=project,
                action=("in" if action == "IN" else "out"),
                device_ts=(timezone.make_aware(datetime.fromisoformat(device_ts))
                           if device_ts else now),
                source="qr",
            )
            duplicate = False
        except IntegrityError:
            # Ya fue recibido antes con el mismo client_uuid
            duplicate = True

        # 2) Actualizar Attendance bajo lock
        att, _ = Attendance.objects.select_for_update().get_or_create(
            project=project, employee=employee, date=today
        )
        att.evidence_photo = photo

        if action == "IN":
            if att.check_in is None:
                att.check_in = now
                att.save()
                msg = f"[{project.name}] Check-In {employee.full_name} {timezone.localtime(now).strftime('%H:%M')}."
            else:
                msg = f"{employee.full_name} ya tenía Check-In {timezone.localtime(att.check_in).strftime('%H:%M')}."
        else:
            if att.check_in is None:
                # Error lógico (no hay IN)
                if request.headers.get("Accept", "").find("application/json") >= 0:
                    return JsonResponse({"ok": False, "error": "No hay Check-In previo."}, status=409)
                messages.error(request, "No hay Check-In previo hoy. Primero registra el ingreso.")
                return redirect("scan", project_slug=project.slug)

            if att.check_out is None:
                att.check_out = now
                att.save()
            # mensaje
            net_mins = _billable_minutes(att)
            net_hhmm = format_hhmm_from_minutes(net_mins)
            is_saturday = (att.date.weekday() == 5)
            note = "" if is_saturday else ", desc. almuerzo 1h"
            msg = (f"[{project.name}] Check-Out {employee.full_name} "
                   f"{timezone.localtime(now).strftime('%H:%M')} (trabajado neto {net_hhmm}{note}).")

        # Respuesta segura: si el cliente pidió JSON, confirmar éxito real
        if request.headers.get("Accept", "").find("application/json") >= 0:
            return JsonResponse({"ok": True, "duplicate": duplicate, "message": msg})

        # Flujos antiguos con mensajes/redirect
        messages.success(request, msg)
        return redirect("scan", project_slug=project.slug)

    # GET
    form = ScanForm(project=project)
    return render(request, "attendance/scan.html", {"form": form, "project": project})

# ──────────────────────────────────────────────────────────────────────────────
# QR por proyecto → genera token firmado y apunta al modo KIOSCO (público)
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def qr_scan_view(request, project_slug):
    project = get_object_or_404(Project, slug=project_slug)
    if not _user_can_access_project(request.user, project):
        raise Http404()

    signer = TimestampSigner(salt="kiosk-scan-v1")
    today_str = timezone.localdate().isoformat()

    # 👇 incluimos la fecha en el payload (cambia cada día)
    payload = f"{project.id}:{project.qr_secret}:{today_str}"
    token = signer.sign(payload)

    kiosk_url = request.build_absolute_uri(
        reverse("scan_kiosk", args=[project.slug]) + "?" + urlencode({"t": token})
    )

    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(kiosk_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    stream = BytesIO()
    img.save(stream, format="PNG")
    return HttpResponse(stream.getvalue(), content_type="image/png")

# ──────────────────────────────────────────────────────────────────────────────
# Scan KIOSCO (público): sin login, sin menús, validando token del QR
# ──────────────────────────────────────────────────────────────────────────────

from django.http import JsonResponse
from django.db import transaction, IntegrityError
import uuid

@csrf_protect
@transaction.atomic
def scan_kiosk_view(request, project_slug):
    project = get_object_or_404(Project, slug=project_slug)

    # ---- Validación del token diario ----
    token = request.GET.get("t")
    if not token:
        return HttpResponse("Token requerido", status=403)

    signer = TimestampSigner(salt="kiosk-scan-v1")
    try:
        data = signer.unsign(token, max_age=36*3600)  # id:secret:YYYY-MM-DD
        parts = data.split(":")
        if len(parts) != 3:
            return HttpResponse("Token inválido", status=403)
        proj_id, secret, token_date = parts
        today_str = timezone.localdate().isoformat()
        if str(project.id) != proj_id or secret != project.qr_secret:
            return HttpResponse("Token inválido", status=403)
        if token_date != today_str:
            return HttpResponse("Token expirado para el día de hoy", status=403)
    except SignatureExpired:
        return HttpResponse("Token expirado", status=403)
    except BadSignature:
        return HttpResponse("Token inválido", status=403)

    # ---- POST (escaneo) ----
    if request.method == "POST":
        form = ScanForm(request.POST, request.FILES, project=project)
        if not form.is_valid():
            if request.headers.get("Accept", "").find("application/json") >= 0:
                return JsonResponse({"ok": False, "error": "Formulario inválido"}, status=400)
            messages.error(request, "Datos inválidos.")
            return redirect(f"{reverse('scan_kiosk', args=[project.slug])}?t={urlquote(token)}")

        employee = form.cleaned_data["employee"]
        action   = form.cleaned_data["action"]
        photo    = form.cleaned_data["evidence_photo"]

        client_uuid = request.POST.get("client_uuid") or str(uuid.uuid4())
        device_ts   = request.POST.get("device_ts")

        today = timezone.localdate()
        now   = timezone.now()

        project.employees.add(employee)

        # 1) Registrar escaneo idempotente
        try:
            from .models import AttendanceScan
            AttendanceScan.objects.create(
                client_uuid=client_uuid,
                employee=employee,
                project=project,
                action=("in" if action == "IN" else "out"),
                device_ts=(timezone.make_aware(datetime.fromisoformat(device_ts))
                           if device_ts else now),
                source="qr",
            )
            duplicate = False
        except IntegrityError:
            duplicate = True  # ya procesado

        # 2) Actualizar Attendance con bloqueo
        att, _ = Attendance.objects.select_for_update().get_or_create(
            project=project, employee=employee, date=today
        )
        att.evidence_photo = photo

        if action == "IN":
            if att.check_in is None:
                att.check_in = now
                att.save()
                msg = f"Check-In {employee.full_name} {timezone.localtime(now).strftime('%H:%M')}."
            else:
                msg = f"{employee.full_name} ya tenía Check-In {timezone.localtime(att.check_in).strftime('%H:%M')}."
        else:
            if att.check_in is None:
                if request.headers.get("Accept", "").find("application/json") >= 0:
                    return JsonResponse({"ok": False, "error": "No hay Check-In previo."}, status=409)
                messages.error(request, "No hay Check-In previo hoy. Primero registra el ingreso.")
                return redirect(f"{reverse('scan_kiosk', args=[project.slug])}?t={urlquote(token)}")

            if att.check_out is None:
                att.check_out = now
                att.save()

            net_mins = _billable_minutes(att)
            net_hhmm = format_hhmm_from_minutes(net_mins)
            is_saturday = (att.date.weekday() == 5)
            note = "" if is_saturday else ", desc. almuerzo 1h"
            msg = f"Check-Out {employee.full_name} {timezone.localtime(now).strftime('%H:%M')} (trabajado neto {net_hhmm}{note})."

        # JSON → confirma éxito real al móvil; el cliente puede reintentar si no recibe esto
        if request.headers.get("Accept", "").find("application/json") >= 0:
            return JsonResponse({"ok": True, "duplicate": duplicate, "message": msg})

        # Flujo tradicional (mensajes + redirect manteniendo token)
        messages.success(request, msg)
        return redirect(f"{reverse('scan_kiosk', args=[project.slug])}?t={urlquote(token)}")

    # ---- GET (formulario kiosco) ----
    form = ScanForm(project=project)
    return render(
        request,
        "attendance/scan_kiosk.html",
        {"form": form, "project": project, "kiosk": True}
    )


# ──────────────────────────────────────────────────────────────────────────────
# Reporte (filtro por mes, nombre y proyecto) + totales y pagos (paginado por día)
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def report_view(request):
    """
    Reporte mensual con filtros:
      - month: YYYY-MM
      - name: coincidencia parcial en nombre
      - project: slug del proyecto

    Admin puede ajustar manualmente los minutos netos (override).
    La lista se pagina por DÍA: página 1 = día más reciente con registros,
    siguientes páginas = días anteriores.
    """
    # ── Admin: procesar ajustes manuales (POST) ───────────────────────────────
    if request.method == "POST" and request.user.is_superuser:
        att_id = request.POST.get("att_id")
        action = request.POST.get("action")
        try:
            att = Attendance.objects.select_related("employee", "project").get(pk=att_id)
        except Attendance.DoesNotExist:
            messages.error(request, "Registro no encontrado.")
            return redirect("report")

        if action == "save":
            try:
                minutes = _parse_hhmm_to_minutes(request.POST.get("hhmm"))
            except Exception:
                messages.error(request, "Formato inválido. Usa HH:MM (ej. 08:00) o horas (ej. 8.5).")
                return redirect("report")

            note = (request.POST.get("note") or "").strip()
            att.manual_minutes = minutes
            att.manual_note = note[:200]
            att.save(update_fields=["manual_minutes", "manual_note"])
            messages.success(
                request,
                f"Horas ajustadas a {minutes//60:02d}:{minutes%60:02d} para {att.employee.full_name}."
            )
        elif action == "clear":
            att.manual_minutes = None
            att.manual_note = ""
            att.save(update_fields=["manual_minutes", "manual_note"])
            messages.success(request, f"Se eliminó el ajuste manual de {att.employee.full_name}.")
        else:
            messages.error(request, "Acción no válida.")
        return redirect("report")

    # ── Filtros GET ───────────────────────────────────────────────────────────
    q_name = (request.GET.get("name") or "").strip()
    q_month_raw = (request.GET.get("month") or "").strip()
    q_proj = (request.GET.get("project") or "").strip()  # slug
    page = max(1, int((request.GET.get("page") or "1").strip() or "1"))

    today = timezone.localdate()
    # Parseo robusto de month
    try:
        if q_month_raw:
            parts = q_month_raw.split("-")
            if len(parts) != 2:
                raise ValueError("month inválido")
            year = int(parts[0])
            month = int(parts[1])
        else:
            year, month = today.year, today.month
    except Exception:
        year, month = today.year, today.month

    # Rango del mes
    first = date(year, month, 1)
    last = (date(year + (month == 12), (month % 12) + 1, 1) - timedelta(days=1))

    qs = Attendance.objects.select_related("employee", "project").filter(date__range=(first, last))

    if not request.user.is_superuser:
        qs = qs.filter(project__in=request.user.supervised_projects.all())

    if q_proj:
        qs = qs.filter(project__slug=q_proj)

    if q_name:
        qs = qs.filter(employee__full_name__icontains=q_name)

    # Obtenemos TODOS los registros del mes (para saber qué días hay)
    all_rows = list(qs.order_by("-date", "project__name", "employee__full_name"))

    # Fechas con datos, ordenadas de más reciente a más antigua
    dates_with_rows = sorted({r.date for r in all_rows}, reverse=True)
    total_pages = max(1, len(dates_with_rows))

    # Ajustar page a rango válido
    if page > total_pages:
        page = total_pages

    # Fecha a mostrar en esta página (si no hay registros, None)
    current_date = dates_with_rows[page - 1] if dates_with_rows else None

    # Filtramos filas SOLO del día actual paginado
    rows = [r for r in all_rows if r.date == current_date] if current_date else []

    # --- Enriquecer filas para la UI (del día actual) ---
    for a in rows:
        # minutos brutos del día (si hay in/out)
        if a.check_in and a.check_out:
            gross = int((a.check_out - a.check_in).total_seconds() // 60)
        else:
            gross = 0

        a.is_saturday = (a.date.weekday() == 5)
        a.lunch_deducted = (not a.is_saturday) and (gross > 300) and a.check_in and a.check_out

        # HH:MM neto con la lógica del modelo (override + reglas)
        a.worked_minutes_net = int(a.worked_minutes())
        a.worked_hhmm_net = f"{a.worked_minutes_net//60:02d}:{a.worked_minutes_net%60:02d}"

    # Totales por (proyecto, empleado) — del día actual
    totals = {}
    grand_minutes = 0
    grand_pay = Decimal("0.00")

    for a in rows:
        key = (a.project.slug, a.employee.full_name)
        minutes = int(a.worked_minutes())  # neto
        rate = a.employee.hourly_rate or Decimal("0.00")

        data = totals.setdefault(
            key,
            {"project": a.project.name, "name": a.employee.full_name, "minutes": 0, "rate": rate, "pay": Decimal("0.00")},
        )
        data["minutes"] += minutes
        data["pay"] = (Decimal(data["minutes"]) / Decimal(60)) * data["rate"]

    # Formateo HH:MM y acumulados
    for info in totals.values():
        m = int(info["minutes"])
        info["hhmm"] = f"{m//60:02d}:{m%60:02d}"
        info["pay_fmt"] = format_money_latam(info["pay"])
        info["rate_fmt"] = format_money_latam(info["rate"])
        grand_minutes += m
        grand_pay += info["pay"]

    grand_hhmm = f"{grand_minutes//60:02d}:{grand_minutes%60:02d}"
    grand_pay = grand_pay.quantize(Decimal("0.01"))
    if not request.user.is_superuser:
        for info in totals.values():
            info["rate"] = None
            info["pay"] = None
            info["rate_fmt"] = None
            info["pay_fmt"] = None
        grand_pay = None
        grand_pay_fmt = None
    else:
        grand_pay_fmt = format_money_latam(grand_pay)

    projects = Project.objects.all() if request.user.is_superuser else request.user.supervised_projects.all()

    # Datos de paginación para el template
    pagination = {
        "page": page,
        "total_pages": total_pages,
        "has_prev": page > 1,
        "has_next": page < total_pages,
        "prev_page": page - 1 if page > 1 else None,
        "next_page": page + 1 if page < total_pages else None,
        "current_date": current_date,  # el día que se está viendo
        "days_count": len(dates_with_rows),
    }

    return render(
        request,
        "attendance/report.html",
        {
            "rows": rows,
            "totals": totals,
            "grand_hhmm": grand_hhmm,
            "grand_pay": grand_pay,
            "grand_pay_fmt": grand_pay_fmt,
            "month": f"{year}-{month:02d}",
            "q_name": q_name,
            "q_proj": q_proj,
            "projects": projects,
            "pagination": pagination,   # ← úsalo para dibujar botones «Anterior / Siguiente»
        },
    )
# ──────────────────────────────────────────────────────────────────────────────
# Empleados
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def employee_list(request):
    """
    Lista de empleados con filtros:
      - project: slug del proyecto (opcional)
      - name: búsqueda por nombre (opcional)

    Si el usuario no es admin, sólo ve empleados que estén en
    sus proyectos supervisados.
    """
    # proyectos disponibles para el filtro (según permisos)
    if request.user.is_superuser:
        projects_qs = Project.objects.all().order_by("name")
        employees = Employee.objects.all()
    else:
        projects_qs = request.user.supervised_projects.all().order_by("name")
        employees = Employee.objects.filter(projects__in=projects_qs).distinct()

    # filtros
    q_proj = (request.GET.get("project") or "").strip()   # slug
    q_name = (request.GET.get("name") or "").strip()

    if q_proj:
        employees = employees.filter(projects__slug=q_proj)

    if q_name:
        employees = employees.filter(full_name__icontains=q_name)

    employees = employees.order_by("full_name").prefetch_related("projects")

    context = {
        "employees": employees,
        "projects": projects_qs,
        "q_proj": q_proj,
        "q_name": q_name,
    }
    return render(request, "attendance/employee_list.html", context)


@login_required
def employee_form(request, pk=None):
    emp = get_object_or_404(Employee, pk=pk) if pk else None
    if request.method == "POST":
        form = EmployeeForm(request.POST, instance=emp)
        if form.is_valid():
            form.save()
            messages.success(request, "Empleado guardado.")
            return redirect("employee_list")
    else:
        form = EmployeeForm(instance=emp)
    return render(request, "attendance/employee_form.html", {"form": form, "emp": emp})


# ──────────────────────────────────────────────────────────────────────────────
# Nómina (solo admin)
# ──────────────────────────────────────────────────────────────────────────────

def _to_decimal(x, default="0"):
    if x is None:
        return Decimal(default)
    if isinstance(x, Decimal):
        return x
    return Decimal(str(x))

TWOPL = Decimal("0.01")

def _monday(d: date) -> date:
    return d - timedelta(days=d.weekday())

def _sunday(d: date) -> date:
    return d + timedelta(days=(6 - d.weekday()))

def _week_chunks(start: date, end: date):
    """Devuelve lista de (week_start, week_end) cubriendo [start, end] en semanas LUN→DOM."""
    cur = _monday(start)
    last = _sunday(end)
    chunks = []
    while cur <= last:
        w_start = cur
        w_end = cur + timedelta(days=6)
        chunks.append((w_start, w_end))
        cur = w_end + timedelta(days=1)
    return chunks

def _build_week_table(employees, att_map, week_start: date, week_end: date, period_start: date, period_end: date):
    """
    Construye la estructura que consume el template de nómina (LUN→DOM).
    att_map: {(emp_id, date): Decimal(hours)}
    """
    cols = [week_start + timedelta(days=i) for i in range(7)]
    col_labels = [d.strftime("%a %d") for d in cols]

    rows = []
    week_hours_all = Decimal("0")
    week_payroll_cent = Decimal("0")

    for e in employees:
        day_hours_display = []
        day_hours_hhmm = []
        total_h = Decimal("0")
        rate = _to_decimal(getattr(e, "hourly_rate", Decimal("0")), "0")

        for d in cols:
            if period_start <= d <= period_end:
                h = _to_decimal(att_map.get((e.id, d), Decimal("0")), "0")
            else:
                h = Decimal("0")

            day_hours_display.append(f"{h.quantize(Decimal('0.00'))}")

            mins = int((h * Decimal("60")).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
            day_hours_hhmm.append(format_hhmm_from_minutes(mins))

            total_h += h

        total_pay_cent = (total_h * rate).quantize(TWOPL, rounding=ROUND_HALF_UP)

        total_h_dec_str = f"{total_h.quantize(Decimal('0.00'))}"
        total_h_mins = int((total_h * Decimal("60")).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
        total_h_hhmm_str = format_hhmm_from_minutes(total_h_mins)

        rows.append({
            "name": e.full_name,
            "day_hours": day_hours_display,        # decimal (si lo quieres)
            "day_hours_hhmm": day_hours_hhmm,      # ← usar en template semanal
            "total_hours": total_h_dec_str,        # decimal (si lo quieres)
            "total_hours_hhmm": total_h_hhmm_str,  # ← usar en template semanal
            "hourly": format_money_latam(rate),
            "total_pay": format_money_latam_whole(total_pay_cent),
        })

        week_hours_all += total_h
        week_payroll_cent += total_pay_cent

    total_hours_all_dec_str = f"{week_hours_all.quantize(Decimal('0.00'))}"
    total_hours_all_mins = int((week_hours_all * Decimal("60")).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    total_hours_all_hhmm_str = format_hhmm_from_minutes(total_hours_all_mins)

    return {
        "columns": col_labels,
        "rows": rows,
        "total_hours_all": total_hours_all_dec_str,
        "total_hours_all_hhmm": total_hours_all_hhmm_str,
        "total_payroll": format_money_latam_whole(week_payroll_cent),
        "total_hours_all_raw": week_hours_all,
        "total_payroll_raw": week_payroll_cent,
        "week_start": week_start,
        "week_end": week_end,
        "show_divider": True,
    }

@login_required
@user_passes_test(is_admin)
def payroll_list(request):
    """
    Lista de periodos agrupados por proyecto (incluye grupo 'Todos' cuando el periodo no tiene proyecto).
    Orden: nombre de proyecto A→Z y, dentro de cada grupo, fecha de inicio descendente.
    """
    qs = (
        PayrollPeriod.objects
        .select_related("project")
        .order_by("project__name", "-start_date", "-id")
    )

    from collections import OrderedDict
    groups = OrderedDict()

    for p in qs:
        key = p.project_id or 0  # 0 para "Todos"
        name = p.project.name if p.project_id else "Todos"
        if key not in groups:
            groups[key] = {
                "name": name,
                "project": p.project,     # puede ser None
                "periods": [],
                "open_count": 0,
                "closed_count": 0,
            }
        groups[key]["periods"].append(p)
        if p.status == "OPEN":
            groups[key]["open_count"] += 1
        else:
            groups[key]["closed_count"] += 1

    # Ordenar alfabéticamente, "Todos" al final
    sorted_groups = sorted(
        groups.values(),
        key=lambda g: (g["name"] == "Todos", g["name"].lower())
    )

    return render(
        request,
        "attendance/payroll_list.html",
        {"groups": sorted_groups}
    )


@login_required
@user_passes_test(is_admin)
def payroll_new(request):
    """Crea una quincena; autopropone 1–15 o 16–fin según hoy."""
    today = timezone.localdate()
    if request.method == "POST":
        form = PayrollPeriodForm(request.POST)
        if form.is_valid():
            period = form.save(commit=False)
            period.created_by = request.user
            period.save()
            messages.success(request, "Periodo creado. Ahora calculamos líneas.")
            return redirect("payroll_recalc", pk=period.pk)
    else:
        y, m = today.year, today.month
        _, last_day = monthrange(y, m)
        if today.day <= 15:
            initial = {"start_date": date(y, m, 1), "end_date": date(y, m, 15)}
        else:
            initial = {"start_date": date(y, m, 16), "end_date": date(y, m, last_day)}
        form = PayrollPeriodForm(initial=initial)
    return render(request, "attendance/payroll_new.html", {"form": form})

@login_required
@user_passes_test(is_admin)
def payroll_recalc(request, pk):
    """Recalcula todas las líneas de la quincena a partir de Attendance (minutos exactos)."""
    period = get_object_or_404(PayrollPeriod, pk=pk)
    if not period.is_open:
        messages.warning(request, "El periodo está cerrado.")
        return redirect("payroll_detail", pk=period.pk)

    qs = Attendance.objects.select_related("employee", "project").filter(
        date__range=(period.start_date, period.end_date)
    )
    if period.project_id:
        qs = qs.filter(project=period.project)

    # minutos NETOS por empleado (–1h almuerzo/override/redondeo)
    minutes_map = defaultdict(int)
    for a in qs:
        minutes_map[a.employee_id] += _billable_minutes(a)

    existing = {l.employee_id: l for l in period.lines.all()}
    period.lines.exclude(employee_id__in=minutes_map.keys()).delete()

    for emp_id, mins in minutes_map.items():
        emp = Employee.objects.get(pk=emp_id)
        line = existing.get(emp_id) or PayrollLine(
            period=period,
            employee=emp,
            adjustment=Decimal("0.00"),
        )
        line.minutes = int(mins)
        hourly = _to_decimal(getattr(emp, "hourly_rate", Decimal("0")), "0")
        line.hourly_rate = hourly
        base_amount = (Decimal(line.minutes) / Decimal(60)) * hourly
        line.base_amount = base_amount.quantize(TWOPL, rounding=ROUND_HALF_UP)  # contable
        line.save()

    messages.success(request, "Nómina recalculada.")
    return redirect("payroll_detail", pk=period.pk)

@login_required
@user_passes_test(is_admin)
def payroll_detail(request, pk):
    """
    Detalle de nómina:
    - Cálculo con minutos exactos (contable a centavos).
    - UI: horas HH:MM y totales a ENTERO.
    - Semanas LUN→DOM con 7 columnas (días fuera del rango = 0).
    """
    period = get_object_or_404(PayrollPeriod, pk=pk)
    lines = list(period.lines.select_related("employee").all())

    # ---- Form para ajuste por línea ----
    line_id = request.GET.get("line")
    form = None
    if line_id:
        line = get_object_or_404(PayrollLine, pk=line_id, period=period)
        if request.method == "POST" and period.is_open:
            form = PayrollLineAdjustmentForm(request.POST, instance=line)
            if form.is_valid():
                form.save()
                messages.success(request, "Ajuste guardado.")
                return redirect("payroll_detail", pk=period.pk)
        else:
            form = PayrollLineAdjustmentForm(instance=line)

    # ---- Totales generales (contables) ----
    grand_base = sum((Decimal(l.base_amount or 0) for l in lines), start=Decimal("0.00"))
    grand_adj  = sum((Decimal(l.adjustment  or 0) for l in lines), start=Decimal("0.00"))
    grand_total = (grand_base + grand_adj).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    # ---- Presentación por línea (UI entero + HH:MM) ----
    lines_fmt = []
    for l in lines:
        mins = int(l.minutes or 0)
        total_line_cent = (Decimal(l.base_amount or 0) + Decimal(l.adjustment or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        lines_fmt.append({
            "obj": l,
            "employee": l.employee,
            "hours_hhmm": format_hhmm_from_minutes(mins),    # HH:MM
            "rate": format_money_latam(l.hourly_rate),
            "base": format_money_latam(l.base_amount),
            "adj":  format_money_latam(l.adjustment),
            "total":      format_money_latam_whole(total_line_cent),  # UI entero
            "total_cent": format_money_latam(total_line_cent),
        })

    # ---- División por semanas (LUN→DOM) ----
    employees = [l.employee for l in lines]
    if employees:
        att_qs = Attendance.objects.filter(
            employee_id__in=[e.id for e in employees],
            date__gte=period.start_date,
            date__lte=period.end_date,
        )
        if period.project_id:
            att_qs = att_qs.filter(project=period.project)

        att_map = defaultdict(lambda: Decimal("0"))
        for a in att_qs:
            mins = _billable_minutes(a)  # neto
            att_map[(a.employee_id, a.date)] += (Decimal(mins) / Decimal(60))
    else:
        att_map = defaultdict(lambda: Decimal("0"))

    week_ranges = _week_chunks(period.start_date, period.end_date)

    week_tables = []
    total_period_pay_cent = Decimal("0")
    total_period_hours = Decimal("0")
    for w_start, w_end in week_ranges:
        wk = _build_week_table(
            employees,
            att_map,
            w_start,
            w_end,
            period.start_date,
            period.end_date,
        )
        week_tables.append(wk)
        total_period_pay_cent += wk["total_payroll_raw"]
        total_period_hours += wk["total_hours_all_raw"]

    context = {
        "period": period,
        "lines": lines,
        "lines_fmt": lines_fmt,
        "form": form,

        # Totales del período (UI entero)
        "grand_base": grand_base,
        "grand_adj": grand_adj,
        "grand_total": grand_total,
        "grand_base_fmt":  format_money_latam_whole(grand_base),
        "grand_adj_fmt":   format_money_latam_whole(grand_adj),
        "grand_total_fmt": format_money_latam_whole(grand_total),

        "title": f"{period.start_date} → {period.end_date}",
        "weeks": week_tables,
        "grand_total_hours": f"{total_period_hours.quantize(Decimal('0.00'))}",
        "grand_total_payroll": format_money_latam_whole(total_period_pay_cent),
    }
    return render(request, "attendance/payroll_detail.html", context)

@login_required
@user_passes_test(is_admin)
def payroll_close(request, pk):
    period = get_object_or_404(PayrollPeriod, pk=pk)
    if period.is_open and request.method == "POST":
        period.status = "CLOSED"
        period.save(update_fields=["status"])
        messages.success(request, "Periodo cerrado.")
    return redirect("payroll_detail", pk=period.pk)

# ──────────────────────────────────────────────────────────────────────────────
# Export Excel de nómina (contable a 2 decimales) — SIN LOGO
# ──────────────────────────────────────────────────────────────────────────────
@login_required
@user_passes_test(is_admin)
def payroll_export_xlsx(request):
    """
    Exporta Nómina a Excel.
    - Si llega ?period_pk=ID → usa líneas ya netas del periodo y genera 2 hojas:
        Hoja 1: Resumen (horas dec., tarifa, base, ajuste, total)
        Hoja 2: Semanas (L→D) como la vista
    - Si llega scope=quincena|mes|anio → agrega por Attendance en el rango.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    CURRENCY_FMT = '"$" #,##0.00'
    HOURS_FMT = '0.00'
    HEADER_FONT = Font(bold=True)
    CENTER = Alignment(horizontal="center")
    RIGHT = Alignment(horizontal="right")

    # ---------- Helpers de formato ----------
    def auto_width(ws, extra=2):
        for col_cells in ws.columns:
            length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + extra, 10), 50)

    def append_header(ws, values):
        ws.append(values)
        for cell in ws[ws.max_row]:
            cell.font = HEADER_FONT
            cell.alignment = CENTER

    def money(cell):
        cell.number_format = CURRENCY_FMT
        cell.alignment = RIGHT

    def hours(cell):
        cell.number_format = HOURS_FMT
        cell.alignment = CENTER

    # ---------- Parámetros ----------
    form = PayrollExportForm(request.GET or None)
    period_pk = request.GET.get("period_pk")

    if not period_pk and not (form.is_bound and form.is_valid()):
        return render(request, "attendance/payroll_export_filter.html", {"form": form})

    project = None
    if form.is_bound and form.is_valid():
        project = form.cleaned_data.get("project")

    # ---------- Caso 1: por period_pk (usa líneas ya netas) ----------
    if period_pk:
        period = get_object_or_404(PayrollPeriod, pk=period_pk)
        lines = list(period.lines.select_related("employee").all())

        wb = Workbook()

        # ================= HOJA 1: RESUMEN =================
        ws = wb.active
        ws.title = "Nómina Periodo"

        ws.append([f"Nómina {period.start_date} → {period.end_date}"])
        ws.append([f"Proyecto: {period.project.name if period.project else 'Todos'}"])
        ws.append([])

        append_header(ws, ["Empleado", "Horas (dec.)", "Tarifa", "Base", "Ajuste", "Total"])

        grand_base = Decimal("0.00")
        grand_adj  = Decimal("0.00")
        grand_tot  = Decimal("0.00")

        for l in sorted(lines, key=lambda x: x.employee.full_name.lower()):
            mins = int(l.minutes or 0)
            hrs  = Decimal(mins) / Decimal(60)
            base = Decimal(l.base_amount or 0)
            adj  = Decimal(l.adjustment  or 0)
            tot  = base + adj

            ws.append([
                l.employee.full_name,
                float(hrs),
                float(Decimal(l.hourly_rate or 0)),
                float(base),
                float(adj),
                float(tot)
            ])
            hours(ws.cell(ws.max_row, 2))
            money(ws.cell(ws.max_row, 3))
            money(ws.cell(ws.max_row, 4))
            money(ws.cell(ws.max_row, 5))
            money(ws.cell(ws.max_row, 6))

            grand_base += base
            grand_adj  += adj
            grand_tot  += tot

        ws.append([])
        ws.append(["TOTALES", "", "", float(grand_base), float(grand_adj), float(grand_tot)])
        money(ws.cell(ws.max_row, 4))
        money(ws.cell(ws.max_row, 5))
        money(ws.cell(ws.max_row, 6))
        ws.cell(ws.max_row, 1).font = HEADER_FONT

        auto_width(ws)

        # ================= HOJA 2: SEMANAS (L→D) =================
        ws2 = wb.create_sheet(title="Semanas (L→D)")

        ws2.append([f"{period.start_date} → {period.end_date}"])
        ws2.append([f"Proyecto: {period.project.name if period.project else 'Todos'}"])
        ws2.append([])

        # Construcción de semanas igual a la vista
        employees = [l.employee for l in lines]
        if employees:
            att_qs = Attendance.objects.filter(
                employee_id__in=[e.id for e in employees],
                date__gte=period.start_date,
                date__lte=period.end_date,
            )
            if period.project_id:
                att_qs = att_qs.filter(project=period.project)

            att_map = defaultdict(lambda: Decimal("0"))
            for a in att_qs:
                mins = _billable_minutes(a)  # neto
                att_map[(a.employee_id, a.date)] += (Decimal(mins) / Decimal(60))
        else:
            att_map = defaultdict(lambda: Decimal("0"))

        week_ranges = _week_chunks(period.start_date, period.end_date)

        total_period_pay_cent = Decimal("0")
        total_period_hours = Decimal("0")

        for idx, (w_start, w_end) in enumerate(week_ranges, start=1):
            wk = _build_week_table(
                employees, att_map, w_start, w_end, period.start_date, period.end_date
            )

            ws2.append([f"Semana {idx}: {w_start} → {w_end}"])

            header = ["NOMBRES Y APELLIDOS"] + wk["columns"] + ["TOTAL HORAS", "VALOR H.", "TOTAL PAGAR"]
            ws2.append(header)
            for c in range(1, len(header) + 1):
                ws2.cell(ws2.max_row, c).font = HEADER_FONT

            for r in wk["rows"]:
                row = [r["name"]] + r["day_hours_hhmm"] + [r["total_hours_hhmm"], r["hourly"], r["total_pay"]]
                ws2.append(row)

            ws2.append(
                ["TOTAL HORAS"] + [""] * len(wk["columns"]) +
                [wk["total_hours_all_hhmm"], "NÓMINA", wk["total_payroll"]]
            )
            for c in range(1, len(header) + 1):
                ws2.cell(ws2.max_row, c).font = HEADER_FONT
            ws2.append([])

            total_period_pay_cent += wk["total_payroll_raw"]
            total_period_hours += wk["total_hours_all_raw"]

        ws2.append(["", "", "", ""])
        ws2.append([
            "TOTAL HORAS (período):",
            f"{total_period_hours.quantize(Decimal('0.00'))}",
            "NÓMINA (período):",
            format_money_latam_whole(total_period_pay_cent),
        ])
        ws2.cell(ws2.max_row, 1).font = HEADER_FONT
        ws2.cell(ws2.max_row, 3).font = HEADER_FONT
        ws2.cell(ws2.max_row, 4).font = HEADER_FONT

        for col_cells in ws2.columns:
            length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
            ws2.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)

        filename = f"nomina_periodo_{period.start_date}_{period.end_date}.xlsx"
        resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(resp)
        return resp

    # ---------- Caso 2: por alcance ----------
    scope = form.cleaned_data["scope"]       # quincena | mes | anio
    month = form.cleaned_data.get("month")
    year = form.cleaned_data.get("year")
    half = form.cleaned_data.get("half")
    today = timezone.localdate()

    if scope == "mes":
        y = (month.year if month else today.year)
        m = (month.month if month else today.month)
        last_day = monthrange(y, m)[1]
        start, end = date(y, m, 1), date(y, m, last_day)
        title = f"Nómina {y}-{m:02d}"
    elif scope == "anio":
        y = year or today.year
        start, end = date(y, 1, 1), date(y, 12, 31)
        title = f"Nómina {y}"
    else:  # quincena
        if not month or not half:
            y, m = today.year, today.month
            last_day = monthrange(y, m)[1]
            if today.day <= 15:
                start, end = date(y, m, 1), date(y, m, 15)
                half = "1"
            else:
                start, end = date(y, m, 16), date(y, m, last_day)
                half = "2"
        else:
            y, m = month.year, month.month
            last_day = monthrange(y, m)[1]
            if half == "1":
                start, end = date(y, m, 1), date(y, m, 15)
            else:
                start, end = date(y, m, 16), date(y, m, last_day)
        title = f"Nómina {y}-{m:02d} Q{half}"

    qs = Attendance.objects.select_related("employee", "project").filter(date__range=(start, end))
    if project:
        qs = qs.filter(project=project)

    wb = Workbook()
    ws = wb.active
    ws.title = "Nomina"

    ws.append([title])
    ws.append([f"Rango: {start} → {end}   |   Proyecto: {project.name if project else 'Todos'}"])
    ws.append([])

    append_header(ws, ["Empleado", "Horas (HH:MM)", "Horas (dec.)", "Tarifa", "Total"])

    # Agregados con minutos NETOS
    agg = defaultdict(lambda: {"minutes": 0, "rate": Decimal("0.00")})
    for a in qs:
        name = a.employee.full_name
        mins = _billable_minutes(a)  # neto
        agg[name]["minutes"] += mins
        agg[name]["rate"] = Decimal(getattr(a.employee, "hourly_rate", 0) or 0)

    grand_total = Decimal("0.00")
    for name, d in sorted(agg.items(), key=lambda x: x[0].lower()):
        hhmm = f"{d['minutes']//60:02d}:{d['minutes']%60:02d}"
        hrs_dec = Decimal(d["minutes"]) / Decimal(60)
        total = (hrs_dec * d["rate"]).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        grand_total += total

        ws.append([name, hhmm, float(hrs_dec), float(d["rate"]), float(total)])
        ws.cell(ws.max_row, 3).number_format = '0.00'
        ws.cell(ws.max_row, 4).number_format = '"$" #,##0.00'
        ws.cell(ws.max_row, 5).number_format = '"$" #,##0.00'

    ws.append([])
    ws.append(["TOTAL", "", "", "", float(grand_total)])
    ws.cell(ws.max_row, 5).number_format = '"$" #,##0.00'
    ws.cell(ws.max_row, 1).font = HEADER_FONT

    # ancho automático
    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 50)

    if scope == "anio":
        filename = f"nomina_anual_{y}_{getattr(project,'slug','todos')}.xlsx"
    elif scope == "mes":
        filename = f"nomina_{y}-{m:02d}_{getattr(project,'slug','todos')}.xlsx"
    else:
        filename = f"nomina_{start}_{end}_{getattr(project,'slug','todos')}.xlsx"

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp

# ──────────────────────────────────────────────────────────────────────────────
# Finanzas de proyecto (crear proyecto, ingresos, gastos y panel)
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_admin)
def project_new(request):
    if request.method == "POST":
        form = ProjectForm(request.POST)
        if form.is_valid():
            p = form.save()
            messages.success(request, "Proyecto creado.")
            return redirect("project_finance_dashboard", project_slug=p.slug)
    else:
        form = ProjectForm()
    return render(request, "attendance/project_new.html", {"form": form})

@login_required
@user_passes_test(is_admin)
def income_new(request):
    if request.method == "POST":
        form = IncomeForm(request.POST)
        if form.is_valid():
            inc = form.save(commit=False)
            inc.created_by = request.user
            inc.save()
            messages.success(request, "Ingreso registrado.")
            return redirect("project_finance_dashboard", project_slug=inc.project.slug)
    else:
        form = IncomeForm()
    return render(request, "attendance/income_form.html", {"form": form})

@login_required
@user_passes_test(is_admin)
def expense_new(request):
    if request.method == "POST":
        form = ExpenseForm(request.POST)
        if form.is_valid():
            exp = form.save(commit=False)
            exp.created_by = request.user
            exp.save()
            messages.success(request, "Gasto registrado.")
            return redirect("project_finance_dashboard", project_slug=exp.project.slug)
    else:
        form = ExpenseForm()
    return render(request, "attendance/expense_form.html", {"form": form})

@login_required
@user_passes_test(is_admin)
def project_finance_dashboard(request, project_slug):
    proj = get_object_or_404(Project, slug=project_slug)

    # --------- Filtros (mes/año) ---------
    today = timezone.localdate()
    month_param = request.GET.get("month")
    year_param = request.GET.get("year")

    if month_param and "-" in month_param:
        y, m = [int(x) for x in month_param.split("-")]
        if year_param:
            y = int(year_param)
    else:
        y = int(year_param) if year_param else today.year
        m = today.month

    last_day = monthrange(y, m)[1]
    start, end = date(y, m, 1), date(y, m, last_day)

    # --------- Agregados ---------
    incomes_qs  = proj.incomes.filter(date__range=(start, end))
    expenses_qs = proj.expenses.filter(date__range=(start, end))

    inc_total = sum((x.amount for x in incomes_qs),  start=Decimal("0.00"))
    exp_total = sum((x.amount for x in expenses_qs), start=Decimal("0.00"))

    att = (
        Attendance.objects
        .filter(project=proj, date__range=(start, end))
        .select_related("employee")
    )
    # costo de nómina con minutos NETOS
    payroll_cost = sum(
        (Decimal(_billable_minutes(a)) / Decimal(60)) * (a.employee.hourly_rate or Decimal("0.00"))
        for a in att
    )
    payroll_cost = (payroll_cost + Decimal("0.00")).quantize(Decimal("0.01"))

    contract_value = proj.contract_value or Decimal("0.00")

    # Balance final
    balance = (contract_value + inc_total) - (exp_total + payroll_cost)

    # Series para gráficos
    step1 = contract_value
    step2 = contract_value + inc_total
    step3 = step2 - exp_total
    step4 = step3 - payroll_cost   # = balance

    line_values = [float(step1), float(step2), float(step3), float(step4), float(step4)]
    chart_values = [
        float(contract_value),
        float(inc_total),
        float(exp_total),
        float(payroll_cost),
        float(balance),
    ]

    # También pasamos strings formateados LATAM para mostrar en la UI
    return render(request, "attendance/project_finance_dashboard.html", {
        "project": proj,
        "year": y, "month": f"{y}-{m:02d}",
        "inc_total": inc_total, "exp_total": exp_total,
        "payroll_cost": payroll_cost, "contract_value": contract_value,
        "balance": balance,
        "inc_total_fmt": format_money_latam(inc_total),
        "exp_total_fmt": format_money_latam(exp_total),
        "payroll_cost_fmt": format_money_latam(payroll_cost),
        "contract_value_fmt": format_money_latam(contract_value),
        "balance_fmt": format_money_latam(balance),
        "incomes": incomes_qs[:20], "expenses": expenses_qs[:20],
        "line_values": line_values,
        "chart_values": chart_values,
    })


# ──────────────────────────────────────────────────────────────────────────────
# Export PDF (A4 horizontal) con MEMBRETE (static/pdf/MEMBRETE.pdf)
# ReportLab para el contenido + PyPDF2 para sobreponer el membrete
# ──────────────────────────────────────────────────────────────────────────────
from io import BytesIO
from decimal import Decimal, ROUND_HALF_UP
from calendar import monthrange
from datetime import date
from collections import defaultdict
import copy

from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils import timezone
from django.contrib.staticfiles import finders

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer

from PyPDF2 import PdfReader, PdfWriter  # pip install pypdf (o PyPDF2)
from reportlab.platypus import PageBreak

@login_required
@user_passes_test(is_admin)
def payroll_export_pdf(request):
    """
    Exporta la nómina a PDF en **horizontal** usando un **membrete** ubicado en:
        static/pdf/MEMBRETE.pdf
    Acepta:
      - ?period_pk=ID   → usa las líneas del período
      - o scope=quincena|mes|anio (+ project opcional)
    """
    # ===== Estilos base
    styles = getSampleStyleSheet()
    H1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=13, leading=15,
                        textColor=colors.HexColor("#dc3545"), spaceAfter=8)
    H2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=11, leading=13,
                        textColor=colors.black, spaceAfter=6)
    SMALL = ParagraphStyle("small", parent=styles["Normal"], fontSize=8.5, leading=10, textColor=colors.gray)

    # ===== Helpers de tabla
    def _build_table(data, col_widths=None, header_bg="#f8d7da", grid="#dc3545",
                     font_size=9, head_font_size=10, align_right_from_col=1):
        t = Table(data, repeatRows=1, colWidths=col_widths)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor(header_bg)),
            ("TEXTCOLOR",  (0,0), (-1,0), colors.black),
            ("GRID",       (0,0), (-1,-1), 0.25, colors.HexColor(grid)),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",   (0,0), (-1,0), head_font_size),
            ("FONTSIZE",   (0,1), (-1,-1), font_size),
            ("ALIGN",      (0,0), (0,-1), "LEFT"),
            ("ALIGN",      (align_right_from_col,1), (-1,-1), "RIGHT"),
        ]))
        return t

    def _col_widths_summary(max_w):
        # Nombre(36%), Horas(10%), Tarifa(18%), Base(18%), Ajuste(9%), Total(9%)
        return [0.36*max_w, 0.10*max_w, 0.18*max_w, 0.18*max_w, 0.09*max_w, 0.09*max_w]

    def _col_widths_for_week(max_w, n_days):
        # Nombre(30%), Días (40% total, partes iguales), TotalH(10%), ValorH(8%), Total(12%)
        name_w = 0.30 * max_w
        days_total = 0.40 * max_w
        day_w = days_total / max(1, n_days)
        rest = [0.10*max_w, 0.08*max_w, 0.12*max_w]
        return [name_w] + [day_w]*n_days + rest

    # ===== Preparar documento (A4 landscape) y story
    form = PayrollExportForm(request.GET or None)
    period_pk = request.GET.get("period_pk")

    content_buf = BytesIO()
    page_size = landscape(A4)

    # Márgenes pensados para que la tabla quepa con membrete horizontal
    # (ajústalos si tu MEMBRETE tiene cabecera/pie altos)
    left = right = 1.2 * 28.3465  # 1.2 cm
    top = bottom = 1.6 * 28.3465  # 1.6 cm (un poco más para no pisar membrete)
    max_table_w = page_size[0] - left - right

    doc = SimpleDocTemplate(
        content_buf,
        pagesize=page_size,
        leftMargin=left, rightMargin=right, topMargin=top, bottomMargin=bottom,
    )
    story = []

    # ───────────── Caso 1: período concreto ─────────────
    if period_pk:
        period = get_object_or_404(PayrollPeriod, pk=period_pk)
        lines = list(period.lines.select_related("employee").all())

        title = f"Nómina {period.start_date} → {period.end_date}"
        subtitle = f"Proyecto: {period.project.name if period.project else 'Todos'}"

        story.append(Paragraph(title, H1))
        story.append(Paragraph(subtitle, SMALL))
        story.append(Spacer(1, 6))

        # Resumen superior
        head = ["Empleado", "Horas (dec.)", "Tarifa", "Base", "Ajuste", "Total"]
        data = [head]
        gb = Decimal("0.00"); ga = Decimal("0.00"); gt = Decimal("0.00")

        for l in sorted(lines, key=lambda x: x.employee.full_name.lower()):
            mins = int(l.minutes or 0)
            hrs  = Decimal(mins) / Decimal(60)
            base = Decimal(l.base_amount or 0)
            adj  = Decimal(l.adjustment  or 0)
            tot  = base + adj
            data.append([
                l.employee.full_name,
                f"{hrs.quantize(Decimal('0.00'))}",
                format_money_latam(l.hourly_rate),
                format_money_latam(base),
                format_money_latam(adj),
                format_money_latam_whole(tot),
            ])
            gb += base; ga += adj; gt += tot

        data.append(["TOTALES", "", "", format_money_latam(grand_base),
             format_money_latam(grand_adj), format_money_latam_whole(grand_tot)])

        # --- Tabla principal (centrada y ancha) ---
        t = Table(
            data,
            repeatRows=1,
            colWidths=[6*cm, 2.5*cm, 2.5*cm, 3*cm, 3*cm, 3*cm],
            hAlign="CENTER"
        )
        t.setStyle(table_header_style())
        story.append(t)
        story.append(Spacer(1, 0.5 * cm))

        # Tabla de Semanas (igual a la vista)
        employees = [l.employee for l in lines]
        if employees:
            att_qs = Attendance.objects.filter(
                employee_id__in=[e.id for e in employees],
                date__gte=period.start_date, date__lte=period.end_date,
            )
            if period.project_id:
                att_qs = att_qs.filter(project=period.project)
            att_map = defaultdict(lambda: Decimal("0"))
            for a in att_qs:
                mins = _billable_minutes(a)
                att_map[(a.employee_id, a.date)] += (Decimal(mins) / Decimal(60))
        else:
            att_map = defaultdict(lambda: Decimal("0"))

        week_ranges = _week_chunks(period.start_date, period.end_date)

        for idx, (w_start, w_end) in enumerate(week_ranges, start=1):
            wk = _build_week_table(
                employees, att_map, w_start, w_end, period.start_date, period.end_date
            )
            story.append(Paragraph(f"Semana {idx}: {w_start} → {w_end}", H2))

            header = ["NOMBRES Y APELLIDOS"] + wk["columns"] + ["TOTAL HORAS", "VALOR H.", "TOTAL PAGAR"]
            rows = [header]
            for r in wk["rows"]:
                rows.append([r["name"]] + r["day_hours_hhmm"] + [r["total_hours_hhmm"], r["hourly"], r["total_pay"]])



# ──────────────────────────────────────────────────────────────────────────────
# Export PDF nómina (Letter Horizontal) + overlay de membrete estático
# Archivo del membrete: static/pdf/MEMBRETE.pdf (en horizontal)
# ──────────────────────────────────────────────────────────────────────────────
from io import BytesIO
from decimal import Decimal, ROUND_HALF_UP
from datetime import date
from calendar import monthrange
from collections import defaultdict

from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render
from django.utils import timezone

from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, PageBreak
)
from reportlab.lib.units import cm

from pypdf import PdfReader, PdfWriter, Transformation

# importa tus helpers ya existentes:
# _billable_minutes, _week_chunks, _build_week_table, format_money_latam, format_money_latam_whole
# Attendance, PayrollPeriod, PayrollExportForm, is_admin


@login_required
@user_passes_test(is_admin)
def payroll_export_pdf(request):
    """
    Exporta la nómina a PDF en formato horizontal y aplica el membrete
    de static/pdf/MEMBRETE.pdf como fondo (encabezado/pie) en cada página.
    Acepta:
      - ?period_pk=ID   (usa líneas del período)
      - o bien ?scope=quincena|mes|anio (+project)
    """
    # ---------------- Estilos ----------------
    styles = getSampleStyleSheet()
    H1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=14,
                        textColor=colors.HexColor("#dc3545"), spaceAfter=6)
    H2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=11,
                        textColor=colors.black, spaceAfter=4)
    SMALL = ParagraphStyle("SMALL", parent=styles["Normal"], fontSize=8,
                           textColor=colors.gray)

    # ---------------- Helpers tabla ----------------
    def table_header_style():
        return TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f8d7da")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dc3545")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 1), (0, -1), "LEFT"),
        ])

    def week_table_style():
        return TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f8d7da")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.black),
            ("GRID",       (0, 0), (-1, -1), 0.25, colors.HexColor("#dc3545")),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN",      (1, 1), (-1, -1), "CENTER"),
            ("ALIGN",      (-3, 1), (-1, -1), "RIGHT"),
        ])

    # ---------------- Construcción de Story (contenido) ----------------
    form = PayrollExportForm(request.GET or None)
    period_pk = request.GET.get("period_pk")

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(letter),
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )
    story = []

    # ── Caso 1: período concreto
    if period_pk:
        period = get_object_or_404(PayrollPeriod, pk=period_pk)
        lines = list(period.lines.select_related("employee").all())

        story.append(Paragraph(f"Nómina {period.start_date} → {period.end_date}", H1))
        story.append(Paragraph(f"Proyecto: {period.project.name if period.project else 'Todos'}", SMALL))
        story.append(Spacer(1, 0.3 * cm))

        # --- Tabla resumen (centrada y ancha) ---
        data = [["Empleado", "Horas (dec.)", "Tarifa", "Base", "Ajuste", "Total"]]
        grand_base = Decimal("0.00")
        grand_adj  = Decimal("0.00")
        grand_tot  = Decimal("0.00")

        for l in sorted(lines, key=lambda x: x.employee.full_name.lower()):
            mins = int(l.minutes or 0)
            hrs  = Decimal(mins) / Decimal(60)
            base = Decimal(l.base_amount or 0)
            adj  = Decimal(l.adjustment  or 0)
            tot  = base + adj
            data.append([
                l.employee.full_name,
                f"{hrs.quantize(Decimal('0.00'))}",
                format_money_latam(l.hourly_rate),
                format_money_latam(base),
                format_money_latam(adj),
                format_money_latam_whole(tot),
            ])
            grand_base += base
            grand_adj  += adj
            grand_tot  += tot

        data.append([
            "TOTALES", "", "",
            format_money_latam(grand_base),
            format_money_latam(grand_adj),
            format_money_latam_whole(grand_tot)
        ])

        t = Table(
            data,
            repeatRows=1,
            colWidths=[6*cm, 2.5*cm, 2.5*cm, 3*cm, 3*cm, 3*cm],
            hAlign="CENTER",
        )
        t.setStyle(table_header_style())
        story.append(t)

        # 👉 Forzar que las semanas empiecen en la 2ª hoja
        story.append(PageBreak())

        # Semanas (como la vista) — cada semana en su propia hoja
        employees = [l.employee for l in lines]
        if employees:
            att_qs = Attendance.objects.filter(
                employee_id__in=[e.id for e in employees],
                date__gte=period.start_date, date__lte=period.end_date,
            )
            if period.project_id:
                att_qs = att_qs.filter(project=period.project)

            att_map = defaultdict(lambda: Decimal("0"))
            for a in att_qs:
                mins = _billable_minutes(a)
                att_map[(a.employee_id, a.date)] += (Decimal(mins) / Decimal(60))
        else:
            att_map = defaultdict(lambda: Decimal("0"))

        week_ranges = _week_chunks(period.start_date, period.end_date)
        for idx, (w_start, w_end) in enumerate(week_ranges, start=1):
            # Salto de página ANTES de cada semana a partir de la 2 (Semana 2, 3, ...)
            if idx > 1:
                story.append(PageBreak())

            wk = _build_week_table(
                employees, att_map, w_start, w_end,
                period.start_date, period.end_date
            )

            story.append(Paragraph(f"Semana {idx}: {w_start} → {w_end}", H2))

            header = ["NOMBRES Y APELLIDOS"] + wk["columns"] + ["TOTAL HORAS", "VALOR H.", "TOTAL PAGAR"]
            rows = [header]
            for r in wk["rows"]:
                rows.append(
                    [r["name"]] + r["day_hours_hhmm"] + [r["total_hours_hhmm"], r["hourly"], r["total_pay"]]
                )
            rows.append(
                ["TOTAL HORAS"] + [""] * len(wk["columns"]) + [wk["total_hours_all_hhmm"], "NÓMINA", wk["total_payroll"]]
            )

            tw = Table(rows, repeatRows=1, hAlign="CENTER")
            tw.setStyle(week_table_style())
            story.append(tw)
            story.append(Spacer(1, 0.4 * cm))

        # Render contenido (sin membrete por ahora)
        doc.build(story)
        content_pdf = buf.getvalue()
        buf.close()

        # Overlay de membrete
        final_pdf = _apply_letter_landscape_template(content_pdf)
        filename = f"nomina_periodo_{period.start_date}_{period.end_date}.pdf"
        resp = HttpResponse(content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        resp.write(final_pdf)
        return resp

    # ── Caso 2: por alcance (quincena / mes / año)
    if not (form.is_bound and form.is_valid()):
        return render(request, "attendance/payroll_export_filter.html", {"form": form})

    project = form.cleaned_data.get("project")
    scope = form.cleaned_data["scope"]
    month = form.cleaned_data.get("month")
    year = form.cleaned_data.get("year")
    half = form.cleaned_data.get("half")
    today = timezone.localdate()

    if scope == "mes":
        y = (month.year if month else today.year)
        m = (month.month if month else today.month)
        last_day = monthrange(y, m)[1]
        start, end = date(y, m, 1), date(y, m, last_day)
        title = f"Nómina {y}-{m:02d}"
    elif scope == "anio":
        y = year or today.year
        start, end = date(y, 1, 1), date(y, 12, 31)
        title = f"Nómina {y}"
    else:
        if not month or not half:
            y, m = today.year, today.month
            last_day = monthrange(y, m)[1]
            start, end = (date(y, m, 1), date(y, m, 15)) if today.day <= 15 else (date(y, m, 16), date(y, m, last_day))
            half = "1" if today.day <= 15 else "2"
        else:
            y, m = month.year, month.month
            last_day = monthrange(y, m)[1]
            start, end = (date(y, m, 1), date(y, m, 15)) if half == "1" else (date(y, m, 16), date(y, m, last_day))
        title = f"Nómina {y}-{m:02d} Q{half}"

    qs = Attendance.objects.select_related("employee", "project").filter(date__range=(start, end))
    if project:
        qs = qs.filter(project=project)

    story.append(Paragraph(title, H1))
    story.append(Paragraph(f"Rango: {start} → {end} — Proyecto: {project.name if project else 'Todos'}", SMALL))
    story.append(Spacer(1, 0.25 * cm))

    # Agregado por empleado (minutos netos)
    agg = defaultdict(lambda: {"minutes": 0, "rate": Decimal("0.00")})
    for a in qs:
        name = a.employee.full_name
        mins = _billable_minutes(a)
        agg[name]["minutes"] += mins
        agg[name]["rate"] = Decimal(getattr(a.employee, "hourly_rate", 0) or 0)

    data = [["Empleado", "Horas (HH:MM)", "Horas (dec.)", "Tarifa", "Total"]]
    grand_total = Decimal("0.00")
    for name, d in sorted(agg.items(), key=lambda x: x[0].lower()):
        hhmm = f"{d['minutes']//60:02d}:{d['minutes']%60:02d}"
        hrs_dec = (Decimal(d["minutes"]) / Decimal(60)).quantize(Decimal("0.00"))
        total = (Decimal(hrs_dec) * d["rate"]).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        grand_total += total
        data.append([name, hhmm, f"{hrs_dec}", format_money_latam(d["rate"]), format_money_latam_whole(total)])

    data.append(["TOTAL", "", "", "", format_money_latam_whole(grand_total)])
    t = Table(
        data,
        repeatRows=1,
        colWidths=[6*cm, 3*cm, 3*cm, 3*cm, 3*cm],
        hAlign="CENTER",
    )
    t.setStyle(table_header_style())
    story.append(t)

    doc.build(story)
    content_pdf = buf.getvalue()
    buf.close()

    final_pdf = _apply_letter_landscape_template(content_pdf)
    if scope == "anio":
        filename = f"nomina_anual_{y}_{getattr(project,'slug','todos')}.pdf"
    elif scope == "mes":
        filename = f"nomina_{y}-{m:02d}_{getattr(project,'slug','todos')}.pdf"
    else:
        filename = f"nomina_{start}_{end}_{getattr(project,'slug','todos')}.pdf"

    resp = HttpResponse(content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp.write(final_pdf)
    return resp


# ---------- Overlay del membrete (LETTER LANDSCAPE) ----------
from django.conf import settings
import os

def _apply_letter_landscape_template(content_pdf_bytes: bytes) -> bytes:
    """
    Toma el PDF de contenido (bytes) y lo coloca dentro del membrete
    estático landscape ubicado en static/pdf/MEMBRETE.pdf.
    - Ajusta el contenido con márgenes y escala si hace falta.
    - Devuelve bytes del PDF final.
    """
    content_reader = PdfReader(BytesIO(content_pdf_bytes))

    template_path = os.path.join(settings.BASE_DIR, "static", "pdf", "MEMBRETE.pdf")
    if not os.path.exists(template_path):
        # Si no hay membrete, devolvemos el contenido tal cual
        return content_pdf_bytes

    template_reader = PdfReader(template_path)
    template_page = template_reader.pages[0]
    tpl_w = float(template_page.mediabox.width)
    tpl_h = float(template_page.mediabox.height)

    writer = PdfWriter()

    # Márgenes dentro del membrete (72 pt = 1 inch)
    left_margin   = 90    # ajusta si necesitas más/menos aire a la izquierda
    right_margin  = 15
    top_margin    = 100   # aire para encabezado
    bottom_margin = 130   # aire extra para pie

    target_inner_w = max(0.0, tpl_w - left_margin - right_margin)
    target_inner_h = max(0.0, tpl_h - top_margin - bottom_margin)

    for page in content_reader.pages:
        c_w = float(page.mediabox.width)
        c_h = float(page.mediabox.height)

        dest = writer.add_blank_page(width=tpl_w, height=tpl_h)

        # Pegar membrete como fondo
        dest.merge_page(template_page)

        # Escala para que el contenido quepa en el área útil
        scale_x = target_inner_w / c_w if c_w else 1.0
        scale_y = target_inner_h / c_h if c_h else 1.0
        scale = min(scale_x, scale_y, 1.0)

        # Trasladar al área útil (márgenes) con esa escala
        t = Transformation().scale(scale).translate(tx=left_margin, ty=bottom_margin)

        dest.merge_transformed_page(page, t)

    out = BytesIO()
    writer.write(out)
    return out.getvalue()
